"""
exporter.py — writes the repository back out in the source workbook's own layout.

Same sheet name, same 27 columns in the same order, so the exported file drops
straight back into the existing reporting process. Two derived columns are
appended at the end because they cannot be recovered from the original layout:
Outcome (Won/Lost, which only ever lived inside Remarks) and the reason a tender
could not be bid.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SHEET = "Tenders"

# (header, record key, column width) — order and spelling match the source file,
# trailing spaces included, so downstream formulas keep working.
COLUMNS: list[tuple[str, str, float]] = [
    ("Sr. No", "srNo", 6),
    ("Organization", "organization", 30),
    ("City", "city", 14),
    ("A/c owner", "owner", 13),
    ("Tender Details ", "details", 42),
    ("Candidate volumes ", "volumeRaw", 14),
    ("Contract period (Years)", "contractPeriod", 11),
    ("Pre Bid meeting date", "preBidDate", 16),
    ("Last date of Pre bid query submission", "queryDate", 16),
    ("Last Date of Bid Submission", "submissionDate", 16),
    ("Tech Opening ", "techOpening", 13),
    ("Tech Presentation", "techPresentation", 14),
    ("Venue visit", "venueVisit", 12),
    ("Commercial Opening ", "commercialOpening", 14),
    ("QCBS", "evaluation", 14),
    ("MSME Pref", "msmePref", 11),
    ("Tender Fees ", "tenderFees", 12),
    ("EMD", "emd", 12),
    ("Status", "status", 17),
    ("Remarks ", "remarks", 30),
    ("Tender Concerns ", "concerns", 34),
    ("PO received", "poReceived", 12),
    ("Agreement signed  (Y/N)", "agreementSigned", 11),
    ("PBG %", "pbg", 8),
    ("Rajan remarks", "rajanRemark", 18),
    ("Chintan remark", "chintanRemark", 18),
    ("Paresh Remark", "pareshRemark", 18),
    # Appended — derived here, absent from the source layout.
    ("Contract Value", "contractValue", 16),
    ("Outcome", "outcome", 18),
    ("Reason not qualified", "noBidReason", 22),
]

DATE_KEYS = {"preBidDate", "queryDate", "submissionDate", "techOpening",
             "techPresentation", "commercialOpening"}
MONEY_KEYS = {"tenderFees", "emd", "contractValue"}

# Back to the workbook's own status vocabulary.
STATUS_OUT = {
    "Submitted": "Submitted",
    "Not Qualified": "not submitted",
    "Under Process": "under process",
    "Cancelled": "Tender getting cancelled",
    "Empanelment": "Empanelment",
    "Unassigned": "",
}

HEAD_FILL = PatternFill("solid", start_color="FF1B3A6B")
HEAD_FONT = Font(bold=True, color="FFFFFFFF", size=10)
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="FFD0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND = PatternFill("solid", start_color="FFF7F8FA")


def _date(value: str):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def build_workbook(rows: list[dict]) -> bytes:
    """Return the repository as .xlsx bytes in the source layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    for i, (header, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, i, header)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 34

    # Sr. No is rewritten sequentially so the export is always 1..n with no gaps.
    ordered = sorted(rows, key=lambda r: (r.get("srNo") or 10**6, r.get("id", "")))

    for r_i, rec in enumerate(ordered, start=2):
        for c_i, (_header, key, _w) in enumerate(COLUMNS, start=1):
            if key == "srNo":
                value = r_i - 1
            elif key == "status":
                value = STATUS_OUT.get(rec.get("category", ""), rec.get("category", ""))
            elif key == "outcome":
                value = "" if rec.get("outcome") in (None, "N/A") else rec["outcome"]
            elif key in DATE_KEYS:
                value = _date(rec.get(key) or "")
            else:
                value = rec.get(key)
                if value is None:
                    value = ""

            cell = ws.cell(r_i, c_i, value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=key in {"details", "remarks", "concerns",
                                                         "rajanRemark", "chintanRemark",
                                                         "pareshRemark"})
            if key in DATE_KEYS and value:
                cell.number_format = "DD-MMM-YYYY"
            elif key in MONEY_KEYS and isinstance(value, (int, float)):
                cell.number_format = '#,##0'
            elif key == "pbg" and isinstance(value, (int, float)):
                cell.number_format = "0%"
            if r_i % 2 == 0:
                cell.fill = BAND

    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(ordered) + 1}"
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
